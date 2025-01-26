import openpyxl

inventory_file = openpyxl.load_workbook("inventory.xlsx")
prod_list = inventory_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
product_under_10 = {}
# Calculation for number of products
for product_row in range(2, prod_list.max_row + 1):
    supplier_name = prod_list.cell(product_row, 4).value
    inventory = prod_list.cell(product_row, 2).value
    price = prod_list.cell(product_row, 3).value
    product_num = prod_list.cell(product_row, 1).value
    inventory_price = prod_list.cell(product_row, 5)

    if supplier_name in product_per_supplier:
        current_num_of_product = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_of_product + 1
    else:
        print("Adding new Supplier")
        product_per_supplier[supplier_name] = 1

    # calculation of total value of per suppliers
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Print inventory which are less than 10
    if inventory < 10:
        product_under_10[int(product_num)] = int(inventory)
    # set value for inventory
    inventory_price.value = inventory * price

    print(product_per_supplier)
    print(total_value_per_supplier)
    print(product_under_10)

inventory_file.save("invent3.xlsx")




